"""
Builds AutoHub_Analytics_Dashboard.xlsx:
  - Raw Data      : full booking-level dataset (with a 'month' helper column)
  - Branch Summary: SUMIFS/COUNTIFS/AVERAGEIFS rollups by branch (formulas, not hardcoded)
  - Category Summary: same, by service category
  - Monthly Trend : same, by month
  - Dashboard     : KPI cards + charts, all formula-driven

Run: python build_excel_dashboard.py
Then: python /mnt/skills/public/xlsx/scripts/recalc.py ../AutoHub_Analytics_Dashboard.xlsx
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

SRC_CSV = "../data/autohub_service_bookings.csv"
OUT_PATH = "../AutoHub_Analytics_Dashboard.xlsx"

# ---------------------------------------------------------------- load & prep
df = pd.read_csv(SRC_CSV, parse_dates=["booking_date"])
df["month"] = df["booking_date"].dt.strftime("%Y-%m")

RAW_COLUMNS = [
    "booking_id", "booking_date", "booking_time", "day_of_week", "branch", "branch_city",
    "service_name", "service_category", "technician", "customer_name", "vehicle_make",
    "vehicle_model", "vehicle_year", "payment_method", "status", "price_pkr",
    "duration_min", "wait_time_min", "customer_rating", "revenue_pkr", "month",
]
df = df[RAW_COLUMNS]

branches = sorted(df["branch"].unique().tolist())
categories = sorted(df["service_category"].unique().tolist())
months = sorted(df["month"].unique().tolist())

N_ROWS = len(df)
LAST_ROW = N_ROWS + 1  # header is row 1

# column letters on Raw Data sheet
COL = {name: get_column_letter(i + 1) for i, name in enumerate(RAW_COLUMNS)}
# booking_id=A booking_date=B booking_time=C day_of_week=D branch=E branch_city=F
# service_name=G service_category=H technician=I customer_name=J vehicle_make=K
# vehicle_model=L vehicle_year=M payment_method=N status=O price_pkr=P
# duration_min=Q wait_time_min=R customer_rating=S revenue_pkr=T month=U

# ---------------------------------------------------------------- styling helpers
HEADER_FILL = PatternFill("solid", fgColor="1A222C")
HEADER_FONT = Font(name="Arial", bold=True, color="EEF2F6", size=11)
TITLE_FONT = Font(name="Arial", bold=True, color="FF5A1F", size=16)
SUBTITLE_FONT = Font(name="Arial", italic=True, color="5D6B7A", size=10)
BODY_FONT = Font(name="Arial", size=10.5, color="10151C")
KPI_LABEL_FONT = Font(name="Arial", size=10, color="5D6B7A")
KPI_VALUE_FONT = Font(name="Arial", bold=True, size=20, color="FF5A1F")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CURRENCY_FMT = '"Rs" #,##0'
PCT_FMT = '0.0%'

def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ================================================================
# SHEET 1: Raw Data
# ================================================================
ws_raw = wb.active
ws_raw.title = "Raw Data"
for j, name in enumerate(RAW_COLUMNS, start=1):
    ws_raw.cell(row=1, column=j, value=name)
style_header_row(ws_raw, 1, len(RAW_COLUMNS))

for i, record in enumerate(df.itertuples(index=False), start=2):
    for j, val in enumerate(record, start=1):
        cell = ws_raw.cell(row=i, column=j, value=val)
        cell.font = BODY_FONT
        if RAW_COLUMNS[j-1] in ("price_pkr", "revenue_pkr"):
            cell.number_format = CURRENCY_FMT
        if RAW_COLUMNS[j-1] == "booking_date":
            cell.number_format = "YYYY-MM-DD"

ws_raw.freeze_panes = "A2"
autofit(ws_raw, [11, 12, 10, 11, 18, 12, 24, 15, 16, 16, 12, 12, 10, 14, 11, 11, 11, 12, 12, 12, 9])

# ================================================================
# SHEET 2: Branch Summary
# ================================================================
ws_b = wb.create_sheet("Branch Summary")
ws_b["A1"] = "Branch Performance Summary"
ws_b["A1"].font = TITLE_FONT
ws_b["A2"] = "All figures computed live via formulas against the Raw Data sheet"
ws_b["A2"].font = SUBTITLE_FONT

headers = ["Branch", "Total Bookings", "Completed Jobs", "Completion Rate", "Revenue (Rs)", "Avg Rating", "Avg Wait (min)"]
for j, h in enumerate(headers, start=1):
    ws_b.cell(row=4, column=j, value=h)
style_header_row(ws_b, 4, len(headers))

def rd_range(col_letter):
    """Full-column data range on Raw Data sheet, e.g. 'Raw Data'!$E$2:$E$6501"""
    return f"'Raw Data'!${col_letter}$2:${col_letter}${LAST_ROW}"

for i, b in enumerate(branches, start=5):
    ws_b.cell(row=i, column=1, value=b).font = BODY_FONT
    ws_b.cell(row=i, column=2, value=f"=COUNTIF({rd_range('E')},A{i})").font = BODY_FONT
    ws_b.cell(row=i, column=3, value=f"=COUNTIFS({rd_range('E')},A{i},{rd_range('O')},\"Completed\")").font = BODY_FONT
    c = ws_b.cell(row=i, column=4, value=f"=C{i}/B{i}")
    c.font = BODY_FONT
    c.number_format = PCT_FMT
    c = ws_b.cell(row=i, column=5, value=f"=SUMIFS({rd_range('T')},{rd_range('E')},A{i},{rd_range('O')},\"Completed\")")
    c.font = BODY_FONT
    c.number_format = CURRENCY_FMT
    c = ws_b.cell(row=i, column=6, value=f"=AVERAGEIFS({rd_range('S')},{rd_range('E')},A{i},{rd_range('O')},\"Completed\")")
    c.font = BODY_FONT
    c.number_format = "0.00"
    c = ws_b.cell(row=i, column=7, value=f"=AVERAGEIFS({rd_range('R')},{rd_range('E')},A{i})")
    c.font = BODY_FONT
    c.number_format = "0.0"
    for col in range(1, 8):
        ws_b.cell(row=i, column=col).border = BORDER

last_b_row = 4 + len(branches)
autofit(ws_b, [20, 15, 15, 16, 16, 12, 15])

# ================================================================
# SHEET 3: Category Summary
# ================================================================
ws_c = wb.create_sheet("Category Summary")
ws_c["A1"] = "Service Category Summary"
ws_c["A1"].font = TITLE_FONT
ws_c["A2"] = "All figures computed live via formulas against the Raw Data sheet"
ws_c["A2"].font = SUBTITLE_FONT

headers2 = ["Category", "Jobs Booked", "Completed Jobs", "Revenue (Rs)", "Avg Price (Rs)", "Avg Rating"]
for j, h in enumerate(headers2, start=1):
    ws_c.cell(row=4, column=j, value=h)
style_header_row(ws_c, 4, len(headers2))

for i, cat in enumerate(categories, start=5):
    ws_c.cell(row=i, column=1, value=cat).font = BODY_FONT
    ws_c.cell(row=i, column=2, value=f"=COUNTIF({rd_range('H')},A{i})").font = BODY_FONT
    ws_c.cell(row=i, column=3, value=f"=COUNTIFS({rd_range('H')},A{i},{rd_range('O')},\"Completed\")").font = BODY_FONT
    c = ws_c.cell(row=i, column=4, value=f"=SUMIFS({rd_range('T')},{rd_range('H')},A{i},{rd_range('O')},\"Completed\")")
    c.font = BODY_FONT
    c.number_format = CURRENCY_FMT
    c = ws_c.cell(row=i, column=5, value=f"=AVERAGEIFS({rd_range('P')},{rd_range('H')},A{i})")
    c.font = BODY_FONT
    c.number_format = CURRENCY_FMT
    c = ws_c.cell(row=i, column=6, value=f"=AVERAGEIFS({rd_range('S')},{rd_range('H')},A{i},{rd_range('O')},\"Completed\")")
    c.font = BODY_FONT
    c.number_format = "0.00"
    for col in range(1, 7):
        ws_c.cell(row=i, column=col).border = BORDER

last_c_row = 4 + len(categories)
autofit(ws_c, [18, 14, 16, 16, 16, 12])

# ================================================================
# SHEET 4: Monthly Trend
# ================================================================
ws_m = wb.create_sheet("Monthly Trend")
ws_m["A1"] = "Monthly Revenue & Volume Trend"
ws_m["A1"].font = TITLE_FONT
ws_m["A2"] = "All figures computed live via formulas against the Raw Data sheet"
ws_m["A2"].font = SUBTITLE_FONT

headers3 = ["Month", "Jobs Booked", "Completed Jobs", "Revenue (Rs)"]
for j, h in enumerate(headers3, start=1):
    ws_m.cell(row=4, column=j, value=h)
style_header_row(ws_m, 4, len(headers3))

for i, mo in enumerate(months, start=5):
    ws_m.cell(row=i, column=1, value=mo).font = BODY_FONT
    ws_m.cell(row=i, column=2, value=f"=COUNTIF({rd_range('U')},A{i})").font = BODY_FONT
    ws_m.cell(row=i, column=3, value=f"=COUNTIFS({rd_range('U')},A{i},{rd_range('O')},\"Completed\")").font = BODY_FONT
    c = ws_m.cell(row=i, column=4, value=f"=SUMIFS({rd_range('T')},{rd_range('U')},A{i},{rd_range('O')},\"Completed\")")
    c.font = BODY_FONT
    c.number_format = CURRENCY_FMT
    for col in range(1, 5):
        ws_m.cell(row=i, column=col).border = BORDER

last_m_row = 4 + len(months)
autofit(ws_m, [12, 14, 16, 16])

# ================================================================
# SHEET 5: Dashboard (KPIs + charts)
# ================================================================
ws_d = wb.create_sheet("Dashboard", 0)  # make it the first sheet
ws_d["B2"] = "AutoHub Vehicle Service Network — Analytics Dashboard"
ws_d["B2"].font = Font(name="Arial", bold=True, size=18, color="EEF2F6")
ws_d["B3"] = "Live formulas pull from Raw Data — refresh by recalculating the workbook"
ws_d["B3"].font = SUBTITLE_FONT

kpi_specs = [
    ("Total Bookings", f"=COUNTA('Raw Data'!$A$2:$A${LAST_ROW})", "0"),
    ("Completed Jobs", f"=COUNTIF('Raw Data'!$O$2:$O${LAST_ROW},\"Completed\")", "0"),
    ("Total Revenue", f"=SUM('Raw Data'!$T$2:$T${LAST_ROW})", CURRENCY_FMT),
    ("Avg Customer Rating", f"=AVERAGE('Raw Data'!$S$2:$S${LAST_ROW})", "0.00"),
    ("Avg Wait Time (min)", f"=AVERAGE('Raw Data'!$R$2:$R${LAST_ROW})", "0.0"),
]

start_col = 2
for i, (label, formula, fmt) in enumerate(kpi_specs):
    col = start_col + i * 2
    cell_label = ws_d.cell(row=5, column=col, value=label)
    cell_label.font = KPI_LABEL_FONT
    cell_val = ws_d.cell(row=6, column=col, value=formula)
    cell_val.font = KPI_VALUE_FONT
    cell_val.number_format = fmt

for row in (5, 6):
    for i in range(len(kpi_specs)):
        col = start_col + i * 2
        ws_d.column_dimensions[get_column_letter(col)].width = 20

# --- Bar chart: Revenue by Branch ---
bar1 = BarChart()
bar1.title = "Revenue by Branch"
bar1.y_axis.title = "Revenue (Rs)"
bar1.style = 10
data = Reference(ws_b, min_col=5, min_row=4, max_row=last_b_row)
cats = Reference(ws_b, min_col=1, min_row=5, max_row=last_b_row)
bar1.add_data(data, titles_from_data=True)
bar1.set_categories(cats)
bar1.height, bar1.width = 9, 15
ws_d.add_chart(bar1, "B9")

# --- Line chart: Monthly revenue trend ---
line1 = LineChart()
line1.title = "Monthly Revenue Trend"
line1.y_axis.title = "Revenue (Rs)"
data = Reference(ws_m, min_col=4, min_row=4, max_row=last_m_row)
cats = Reference(ws_m, min_col=1, min_row=5, max_row=last_m_row)
line1.add_data(data, titles_from_data=True)
line1.set_categories(cats)
line1.height, line1.width = 9, 15
ws_d.add_chart(line1, "J9")

# --- Pie chart: Revenue share by category ---
pie1 = PieChart()
pie1.title = "Revenue Share by Category"
data = Reference(ws_c, min_col=4, min_row=4, max_row=last_c_row)
cats = Reference(ws_c, min_col=1, min_row=5, max_row=last_c_row)
pie1.add_data(data, titles_from_data=True)
pie1.set_categories(cats)
pie1.dataLabels = DataLabelList()
pie1.dataLabels.showPercent = True
pie1.height, pie1.width = 9, 15
ws_d.add_chart(pie1, "B25")

# --- Bar chart: Jobs by branch (volume) ---
bar2 = BarChart()
bar2.title = "Job Volume by Branch"
bar2.y_axis.title = "Jobs"
data = Reference(ws_b, min_col=2, min_row=4, max_row=last_b_row)
cats = Reference(ws_b, min_col=1, min_row=5, max_row=last_b_row)
bar2.add_data(data, titles_from_data=True)
bar2.set_categories(cats)
bar2.height, bar2.width = 9, 15
ws_d.add_chart(bar2, "J25")

ws_d.sheet_view.showGridLines = False

wb.save(OUT_PATH)
print(f"Saved workbook to {OUT_PATH}")
print(f"Branches: {branches}")
print(f"Categories: {categories}")
print(f"Months: {months}")
